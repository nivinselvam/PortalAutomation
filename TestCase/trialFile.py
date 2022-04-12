import time

import openpyxl
import ExcelProcessor
from openpyxl.styles import Font, PatternFill, Border, Side

from TestCase import Rerun
from Utilities import configReader
from datetime import datetime


def changeValueForRerun():
    setOfRerunTest = set()
    setOfRerunTest.add("apple")
    # setOfRerunTest.add("orange")
    # setOfRerunTest.add("grapes")
    # setOfRerunTest.add("banana")
    ls_TestCasesForRerun = list(setOfRerunTest)
    print(ls_TestCasesForRerun)
    print("")

    # print(ls_TestCasesForRerun[2])
    # print(ls_TestCasesForRerun[0])
    # print(ls_TestCasesForRerun[3])
    # print(ls_TestCasesForRerun[1])

    print(len(ls_TestCasesForRerun))


    wb = openpyxl.load_workbook("/home/ezetap/PycharmProjects/PortalAutomation/TestCase/Report.xlsx")
    sheet = wb['Sheet1']

    for tc in ls_TestCasesForRerun:
        for i in range(2, sheet.max_row + 1):
            colNum_testcase = ExcelProcessor.getColumnNumberFromName("", sheet, 'Test Case ID')
            testcase = (sheet.cell(row=i, column=colNum_testcase)).value

            if testcase == tc:
                colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
                sheet.cell(row=i, column=colNum_overallStatus).value = ""
    wb.save("/home/ezetap/PycharmProjects/PortalAutomation/TestCase/Report.xlsx")



def updateExcelWithValues():
    wb = openpyxl.load_workbook("/home/ezetap/PycharmProjects/PortalAutomation/TestCase/Report.xlsx")
    sheet = wb['Sheet1']
    c3 =sheet.cell(row=2,column=1)
    testcase = c3.value

    list_deselected_testcases = ["test_deselected","test_onlyAPI_Val"]

    if "test_success" in list_deselected_testcases:
        print("CONTAINS")

    for i in range(2,sheet.max_row+1):
        colNum_testcase = ExcelProcessor.getColumnNumberFromName("", sheet, 'Test Case ID')
        testcase = (sheet.cell(row=i, column=colNum_testcase)).value

        if testcase in list_deselected_testcases:

            colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
            sheet.cell(row=i, column=colNum_overallStatus).value = "Deselected"

            colNum_execution = ExcelProcessor.getColumnNumberFromName("", sheet, 'TC Execution')
            sheet.cell(row=i, column=colNum_execution).value = "N/A"

            colNum_APIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'API Val')
            sheet.cell(row=i, column=colNum_APIval).value = "N/A"

            colNum_DBval = ExcelProcessor.getColumnNumberFromName("", sheet, 'DB Val')
            sheet.cell(row=i, column=colNum_DBval).value = "N/A"

            colNum_PortalVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'Portal Val')
            sheet.cell(row=i, column=colNum_PortalVal).value = "N/A"

            colNum_AppVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'App Val')
            sheet.cell(row=i, column=colNum_AppVal).value = "N/A"

            colNum_UIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'UI Val')
            sheet.cell(row=i, column=colNum_UIval).value = "N/A"


        else:

            colNum_overallResult = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
            cellValue = (sheet.cell(row=i, column=colNum_overallResult)).value
            if cellValue is None:
                colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
                sheet.cell(row=i, column=colNum_overallStatus).value = "Broken"

                colNum_execution = ExcelProcessor.getColumnNumberFromName("", sheet, 'TC Execution')
                sheet.cell(row=i, column=colNum_execution).value = "N/A"

                colNum_APIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'API Val')
                sheet.cell(row=i, column=colNum_APIval).value = "N/A"

                colNum_DBval = ExcelProcessor.getColumnNumberFromName("", sheet, 'DB Val')
                sheet.cell(row=i, column=colNum_DBval).value = "N/A"

                colNum_PortalVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'Portal Val')
                sheet.cell(row=i, column=colNum_PortalVal).value = "N/A"

                colNum_AppVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'App Val')
                sheet.cell(row=i, column=colNum_AppVal).value = "N/A"

                colNum_UIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'UI Val')
                sheet.cell(row=i, column=colNum_UIval).value = "N/A"

    max_row = sheet.max_row
    max_column = sheet.max_column




    side = Side(border_style="thin", color="000000")

    border = Border(left=side, right=side, top=side, bottom=side)
    for column in range(1, max_column + 1):
        for row in range(1, max_row + 1):
            sheet.cell(row, column).border = border

    fill_pattern = PatternFill(patternType='solid', fgColor='87CEEB')
    font = Font(size=11, bold=True, color="121103")

    print(sheet.max_column)

    for x in range(1, sheet.max_column):
        sheet.cell(row=1, column=x).font = font
        sheet.cell(row=1, column=x).fill = fill_pattern

    wb.save("/home/ezetap/PycharmProjects/PortalAutomation/TestCase/Report.xlsx")



def updateExcel_With_RerunAttempts():
    wb = openpyxl.load_workbook("/home/ezetap/PycharmProjects/PortalAutomation/TestCase/Report.xlsx")
    sheet = wb['Sheet1']

    if configReader.read_config("Validations", "rerun_immediately").lower() == "true" and Rerun.isRerunRequiredImmediately("test_api_common_05"):
        colNum_rerunAttempts = ExcelProcessor.getColumnNumberFromName("", sheet, 'Rerun Attempts')
        for i in range(2, sheet.max_row + 1):
            colNum_testcase = ExcelProcessor.getColumnNumberFromName("", sheet, 'Test Case ID')
            testcase = (sheet.cell(row=i, column=colNum_testcase)).value

            rerunCount = Rerun.getRerunCount(testcase)
            if rerunCount >= 0:
                cellValue_rerunAttempts = configReader.read_config("Validations", "rerun_count") - rerunCount


def change():
    current = datetime.now()
    time_current_one = current.strftime("%H:%M:%S:%f")
    # time.sleep(1)

    current = datetime.now()
    time_current_two = current.strftime("%H:%M:%S:%f")
    print("time_current_one : ",time_current_one)
    print("time_current_two : ",time_current_two)


    diff = datetime.strptime(time_current_two) - (datetime.strptime(time_current_one))


    totalExecutionTime = datetime.strptime(GlobalVariables.EXCEL_TC_Exe_completed_time, FMT) - datetime.strptime(str(
        GlobalVariables.EXCEL_TC_Exe_Starting_Time), FMT)


    print("diff : ",diff)



change()