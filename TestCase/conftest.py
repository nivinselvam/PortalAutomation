import datetime
import os
import openpyxl
from selenium.webdriver.chrome import webdriver
from allure_commons.types import AttachmentType
import allure
import paramiko
import pytest
from selenium import webdriver
from datetime import datetime
import chromedriver_autoinstaller
from random import randint
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Border, Side
from TestCase import ExcelProcessor
from TestCase import Rerun

from Utilities import configReader
from TestCase import server_logs, setUp
from DataProvider import GlobalVariables
from Pages import BaseActions
from ConfigurationData import  Configuration
from TestCase.server_logs import env

executed_tests_count = 0
passed1 = 0
failed1 = 0
skipped1 = 0
deselect1 = 0
rerunCount = 0
LogCollTime = 00
list_deselected_testcases = []
currentTestCase = ""
now = datetime.now()
starting_time = now.strftime("%H:%M:%S")

# router_ip = '192.168.3.73'    #dev3
router_ip = '192.168.3.81'    #dev11
router_username = 'divyaandrews' # Replace with your username in config file
router_port = 22
key_filename = '/home/ezetap/.ssh/divya' # Replace with your private key filename
ssh = paramiko.SSHClient()


def pytest_deselected(items):
    global list_deselected_testcases
    for item in items:
        testcase = str((item.nodeid)).split('::')[1]
        list_deselected_testcases.append(testcase)


@pytest.mark.hookwrapper
@pytest.hookimpl(hookwrapper=True, tryfirst=True)
def pytest_runtest_makereport(item, call):
    pytest_html = item.config.pluginmanager.getplugin('html')
    outcome = yield
    report = outcome.get_result()
    setattr(report, "duration_formatter", "%S")   #"%H:%M:%S.%f"
    setattr(item, "rep_" + report.when, report)
    report.description = str(item.function.__doc__)
    return report


def revert_excel_global_variables():
    GlobalVariables.EXCEL_TC_Exe_Starting_Time = 00
    GlobalVariables.EXCEL_TC_Exe_completed_time = 00

    GlobalVariables.EXCEL_TC_Val_Starting_Time = 00
    GlobalVariables.EXCEL_TC_Val_completed_time = 00

    GlobalVariables.EXCEL_TC_LogColl_Starting_Time = 00
    GlobalVariables.EXCEL_TC_LogColl_completed_time = 00
    GlobalVariables.EXCEL_TC_Execution = "Skip"
    GlobalVariables.EXCEL_API_Val = "N/A"
    GlobalVariables.EXCEL_DB_Val = "N/A"
    GlobalVariables.EXCEL_Portal_Val = "N/A"
    GlobalVariables.EXCEL_App_Val = "N/A"
    GlobalVariables.EXCEL_UI_Val = "N/A"


def setStylesForExcel():
    wb = openpyxl.load_workbook(GlobalVariables.EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    max_row = sheet.max_row

    colNum_overall = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
    colNum_execution = ExcelProcessor.getColumnNumberFromName("", sheet, 'TC Execution')
    colNum_apiVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'API Val')
    colNum_dbVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'DB Val')
    colNum_portalVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'Portal Val')
    colNum_appVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'App Val')
    colNum_uiVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'UI Val')

    column_list = [colNum_overall, colNum_execution, colNum_apiVal, colNum_dbVal, colNum_portalVal, colNum_appVal, colNum_uiVal]

    for column in column_list:
        for row in range(2, max_row + 1):
            if sheet.cell(row, column).value == "Pass":
                sheet.cell(row, column).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif sheet.cell(row, column).value == "Fail":
                sheet.cell(row, column).fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')
            elif sheet.cell(row, column).value == "N/A":
                sheet.cell(row, column).fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    for i in range(2, sheet.max_row + 1):
        colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
        if sheet.cell(row=i, column=colNum_overallStatus).value == "Broken":
            sheet.cell(row=i, column=colNum_overallStatus).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        if sheet.cell(row=i, column=colNum_overallStatus).value == "Deselected":
            sheet.cell(row=i, column=colNum_overallStatus).fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')



    # Set width for all cells with all columns
    sheet.column_dimensions['A'].width = 70 #Test Case ID
    sheet.column_dimensions['B'].width = 23 #File Name
    sheet.column_dimensions['C'].width = 23 #Directory Name
    sheet.column_dimensions['D'].width = 20 #Category
    sheet.column_dimensions['E'].width = 15 #Sub-Category
    sheet.column_dimensions['F'].width = 18 #OverAll Results
    sheet.column_dimensions['G'].width = 15 #TC Execution
    sheet.column_dimensions['H'].width = 15 #API Val
    sheet.column_dimensions['I'].width = 15 #DB Val
    sheet.column_dimensions['J'].width = 15 #Portal Val
    sheet.column_dimensions['K'].width = 15 #App Val
    sheet.column_dimensions['L'].width = 15 #UI Val
    sheet.column_dimensions['M'].width = 22 #Execution Time
    sheet.column_dimensions['N'].width = 25 #Validation Time
    sheet.column_dimensions['O'].width = 25 #Log Coll Time
    sheet.column_dimensions['P'].width = 20 #Total Time
    sheet.column_dimensions['Q'].width = 20 #Rerun Attempts


    # Set background color and font style
    fill_pattern = PatternFill(patternType='solid', fgColor='87CEEB')
    font = Font(size=11, bold=True, color="121103")

    for x in range(1, sheet.max_column+1):
        sheet.cell(row=1, column=x).font = font
        sheet.cell(row=1, column=x).fill = fill_pattern

    # Set border for all the cells
    side = Side(border_style="thin", color="000000")

    border = Border(left=side, right=side, top=side, bottom=side)
    for column in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row, column).border = border

    wb.save(GlobalVariables.EXCEL_reportFilePath)


# WITH DF
@pytest.fixture(scope="function") # Executing once before every testcases
def method_setup(request):
    global LogCollTime
    current = datetime.now()
    LogColl_Starting_Time = current.strftime("%H:%M:%S")

    if env.__contains__('dev'):
        if BaseActions.enter_data_logs("fetch_api_Logs") == "True":
            GlobalVariables.startLineNumberAPI = server_logs.noOfLine(BaseActions.pathToLogFile('api_dev'))
        if BaseActions.enter_data_logs("fetch_portal_Logs") == "True":
            GlobalVariables.startLineNumberPortal = server_logs.noOfLine(BaseActions.pathToLogFile('portal_dev'))
        if BaseActions.enter_data_logs("fetch_middleware_Logs") == "True":
            GlobalVariables.startLineNumberMiddlewware = server_logs.noOfLine(
                BaseActions.pathToLogFile('middleware_dev'))
        if BaseActions.enter_data_logs("fetch_cnpware_Logs") == "True":
            GlobalVariables.startLineNumberCnpware = server_logs.noOfLine(BaseActions.pathToLogFile('cnpware_dev'))
    else:
        if BaseActions.enter_data_logs("fetch_api_Logs") == "True":
            GlobalVariables.startLineNumberAPI = server_logs.noOfLine(BaseActions.pathToLogFile('api'))
        if BaseActions.enter_data_logs("fetch_portal_Logs") == "True":
            GlobalVariables.startLineNumberPortal = server_logs.noOfLine(BaseActions.pathToLogFile('portal'))
        if BaseActions.enter_data_logs("fetch_middleware_Logs") == "True":
            GlobalVariables.startLineNumberMiddlewware = server_logs.noOfLine(BaseActions.pathToLogFile('middleware'))
        if BaseActions.enter_data_logs("fetch_cnpware_Logs") == "True":
            GlobalVariables.startLineNumberCnpware = server_logs.noOfLine(BaseActions.pathToLogFile('cnpware'))

    current = datetime.now()
    LogColl_Ending_Time = current.strftime("%H:%M:%S")
    FMT = '%H:%M:%S'
    totalLogCollectionTime = datetime.strptime(LogColl_Ending_Time, FMT) - datetime.strptime(str(LogColl_Starting_Time),
                                                                                             FMT)

    # Converting time duration to seconds
    LogCollTime = sum(x * int(t) for x, t in zip([3600, 60, 1], str(totalLogCollectionTime).split(":")))

    current = datetime.now()
    GlobalVariables.EXCEL_TC_Exe_Starting_Time = current.strftime("%H:%M:%S")

    # Executing once AFTER every testcases
    def fin():
        GlobalVariables.EXCEL_testCaseName = os.environ.get('PYTEST_CURRENT_TEST').replace(" (teardown)", '').split('::')[1]
        if Rerun.getRerunCountForAtTheEnd() == 0:
            log_on_failure(request)

        writeDataToReportExcelAfterEachTestCase()

        if configReader.read_config("Validations",
                                    "rerun_immediately").lower() == "true" and Rerun.isRerunRequiredImmediately(GlobalVariables.EXCEL_testCaseName)\
                and configReader.read_config("Validations", "rerun_at_the_end").lower() == "false":
            rerunCount = Rerun.getRerunCount(GlobalVariables.EXCEL_testCaseName)
            if rerunCount >= 0:
                print(str(rerunCount) + " reruns pending for the test case " + GlobalVariables.EXCEL_testCaseName)
                rerunCount -= 1
                Rerun.rerunTestImmediately(GlobalVariables.EXCEL_testCaseName, GlobalVariables.EXCEL_testCaseFileName,
                                           rerunCount, request)
            else:
                print(str(rerunCount) + " reruns pending for the test case " + GlobalVariables.EXCEL_testCaseName)
                print("Rerun skipped")


        # Reverting all the global variables back to default values
        revert_excel_global_variables()

    request.addfinalizer(fin)


def writeDataToReportExcelAfterEachTestCase():
    global LogCollTime
    GlobalVariables.EXCEL_LogCollTime = LogCollTime + GlobalVariables.EXCEL_LogCollTime
    GlobalVariables.EXCEL_Tot_Time = int(GlobalVariables.EXCEL_Execution_Time) + int(
        GlobalVariables.EXCEL_Val_time) + int(GlobalVariables.EXCEL_LogCollTime)

    Overall_Status = 'Broken'
    if GlobalVariables.EXCEL_TC_Execution == 'Pass':

        # Pass or N/A
        if GlobalVariables.EXCEL_API_Val != 'Fail' and GlobalVariables.EXCEL_DB_Val != 'Fail' and GlobalVariables.EXCEL_Portal_Val != 'Fail' \
                and GlobalVariables.EXCEL_UI_Val != 'Fail' and GlobalVariables.EXCEL_App_Val !='Fail':
            Overall_Status = 'Pass'

        elif GlobalVariables.EXCEL_API_Val == 'Fail' or GlobalVariables.EXCEL_DB_Val == 'Fail' or GlobalVariables.EXCEL_Portal_Val == 'Fail' \
                or GlobalVariables.EXCEL_UI_Val == 'Fail' or GlobalVariables.EXCEL_App_Val == 'Fail':
            Overall_Status = 'Fail'
    elif GlobalVariables.EXCEL_TC_Execution == 'Fail':
        Overall_Status = 'Fail'

    workbook = openpyxl.load_workbook(GlobalVariables.EXCEL_reportFilePath)
    sheet = workbook["Sheet1"]

    rowNumber = ExcelProcessor.getRowNumberFromValue(workbook, sheet, 'Test Case ID',
                                                     GlobalVariables.EXCEL_testCaseName)

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'File Name')
    GlobalVariables.EXCEL_testCaseFileName = sheet.cell(row=rowNumber, column=columnNumber).value

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'OverAll Results')
    sheet.cell(row=rowNumber, column=columnNumber).value = Overall_Status

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'TC Execution')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_TC_Execution

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'API Val')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_API_Val

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'DB Val')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_DB_Val

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Portal Val')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_Portal_Val

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'App Val')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_App_Val

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'UI Val')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_UI_Val

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Execution Time (sec)')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_Execution_Time

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Validation Time (sec)')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_Val_time

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Log Coll Time (sec)')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_LogCollTime

    columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Total Time (sec)')
    sheet.cell(row=rowNumber, column=columnNumber).value = GlobalVariables.EXCEL_Tot_Time



    # To add the rerun count after every executed testcases
    # if rerun_at_the_end & rerun_immediately are enabled, rerun_at_the_end will be considered
    # if (rerun_at_the_end is TRUE) OR (rerun_at_the_end & rerun_immediately are TRUE)
    if (configReader.read_config("Validations", "rerun_at_the_end").lower() == "true" and configReader.read_config("Validations", "rerun_immediately").lower() == "false") \
            or (configReader.read_config("Validations", "rerun_at_the_end").lower() == "true" and configReader.read_config("Validations", "rerun_immediately").lower() == "true"):
        columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Rerun Attempts')

        if sheet.cell(row=rowNumber, column=columnNumber).value is None:
            sheet.cell(row=rowNumber, column=columnNumber).value = 0
        else:

            currentRetryCountsheet = sheet.cell(row=rowNumber, column=columnNumber).value
            sheet.cell(row=rowNumber, column=columnNumber).value = currentRetryCountsheet + 1

    workbook.save(GlobalVariables.EXCEL_reportFilePath)
    workbook.close()


@pytest.fixture(scope="function")
def ui_driver(request):
    GlobalVariables.portalDriver = chromedriver_autoinstaller.install()
    # Chrome options
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument('--disable-dev-shm-usage')
    # Run chrome
    GlobalVariables.portalDriver = webdriver.Chrome(options=chrome_options)
    GlobalVariables.portalDriver.maximize_window()


@pytest.fixture(scope="function")
def appium_driver():
    desired_cap = {
        "platformName": "Android",
        "deviceName": "1490329804",
        "udid": "1490329804",
        "appPackage": "com.ezetap.basicapp",
        "appActivity": "com.ezetap.mpos.SplashScreen",
        "ignoreHiddenApiPolicyError": "true",
        "noReset": "true",
        "autoGrantPermissions": "true",
        "MobileCapabilityType.AUTOMATION_NAME": "AutomationName.ANDROID_UIAUTOMATOR2"
    }
    GlobalVariables.appDriver = webdriver.Remote("http://localhost:4723/wd/hub", desired_cap)
    GlobalVariables.appDriver.implicitly_wait(30)


def log_on_failure(request):
    yield
    item = request.node
    if item.rep_call.failed:
        current = datetime.now()
        GlobalVariables.EXCEL_TC_LogColl_Starting_Time = current.strftime("%H:%M:%S")

        if BaseActions.enter_data_logs("For_Failed_TCS_fetch_Logs") == "True":
            path = "/home/ezetap/PycharmProjects/PortalAutomation/ServerLogs/" + str(
                item.nodeid.replace("::", "_")).replace("TestCase", "")
            Path(path).mkdir(parents=True, exist_ok=True)
            if GlobalVariables.apiLogs and BaseActions.enter_data_logs("fetch_api_Logs") == "True":
                apiLogs = server_logs.fetchAPILogs()
                server_logs.appendLogs(path + "/api.txt", str(item.nodeid), apiLogs)

            if GlobalVariables.portalLogs and BaseActions.enter_data_logs("fetch_portal_Logs") == "True":
                portalLogs = server_logs.fetchPortalLogs()
                server_logs.appendLogs(path + "/portal.txt", str(item.nodeid), portalLogs)

            if GlobalVariables.middleWareLogs and BaseActions.enter_data_logs("fetch_middleware_Logs") == "True":
                mwareLogs = server_logs.fetchMiddlewareLogs()
                server_logs.appendLogs(path + "/middleware.txt", str(item.nodeid), mwareLogs)

            if GlobalVariables.cnpWareLogs and BaseActions.enter_data_logs("fetch_cnpware_Logs") == "True":
                cnpWareLogs = server_logs.fetchCnpwareLogs()
                server_logs.appendLogs(path + "/cnpware.txt", str(item.nodeid), cnpWareLogs)

        setUp.get_Log_Collection_Time()

        if GlobalVariables.successApp == 'Failed' and GlobalVariables.appDriver != '':
            allure.attach(GlobalVariables.appDriver.get_screenshot_as_png(), name="screenshot",
                          attachment_type=AttachmentType.PNG)
            GlobalVariables.successApp = 'Passed'

        if GlobalVariables.successPortal == 'Failed' and GlobalVariables.portalDriver != '':
            allure.attach(GlobalVariables.portalDriver.get_screenshot_as_png(), name="screenshot",
                          attachment_type=AttachmentType.PNG)
            GlobalVariables.successPortal = 'Passed'

        if GlobalVariables.portalDriver != '':
            GlobalVariables.portalDriver.quit()
            GlobalVariables.portalDriver = ''
        if GlobalVariables.appDriver != '':
            GlobalVariables.appDriver.quit()
            GlobalVariables.appDriver = ''

        GlobalVariables.successPortal = "N/A"
        GlobalVariables.successApp = "N/A"


@pytest.fixture(scope='function')
def log_on_success(request):
    yield
    item = request.node
    if item.rep_call.passed:
        current = datetime.now()
        GlobalVariables.EXCEL_TC_LogColl_Starting_Time = current.strftime("%H:%M:%S")

        path = "/home/ezetap/PycharmProjects/PortalAutomation/ServerLogs/" + str(
            item.nodeid.replace("::", "_")).replace("TestCase", "")
        Path(path).mkdir(parents=True, exist_ok=True)
        if BaseActions.enter_data_logs("For_Passed_TCS_fetch_Logs") == "True":
            if GlobalVariables.apiLogs and BaseActions.enter_data_logs("fetch_api_Logs") == "True":
                apiLogs = server_logs.fetchAPILogs()
                server_logs.appendLogs(path + "/api.txt", str(item.nodeid), apiLogs)

            if GlobalVariables.portalLogs and BaseActions.enter_data_logs("fetch_portal_Logs") == "True":
                portalLogs = server_logs.fetchPortalLogs()
                server_logs.appendLogs(path + "/portal.txt", str(item.nodeid), portalLogs)

            if GlobalVariables.middleWareLogs and BaseActions.enter_data_logs("fetch_middleware_Logs") == "True":
                mwareLogs = server_logs.fetchMiddlewareLogs()
                server_logs.appendLogs(path + "/middleware.txt", str(item.nodeid), mwareLogs)

            if GlobalVariables.cnpWareLogs and BaseActions.enter_data_logs("fetch_cnpware_Logs") == "True":
                cnpWareLogs = server_logs.fetchCnpwareLogs()
                server_logs.appendLogs(path + "/cnpware.txt", str(item.nodeid), cnpWareLogs)

        setUp.get_Log_Collection_Time()

        if BaseActions.enter_data_logs("fetch_ss") == "True" and GlobalVariables.appDriver != '':
            allure.attach(GlobalVariables.appDriver.get_screenshot_as_png(), name="screenshot",
                          attachment_type=AttachmentType.PNG)
        if BaseActions.enter_data_logs("fetch_ss") == "True" and GlobalVariables.portalDriver != '':
            allure.attach(GlobalVariables.portalDriver.get_screenshot_as_png(), name="screenshot",
                          attachment_type=AttachmentType.PNG)

        if GlobalVariables.portalDriver != '':
            GlobalVariables.portalDriver.quit()
            GlobalVariables.portalDriver = ''

        if GlobalVariables.appDriver != '':
            GlobalVariables.appDriver.quit()
            GlobalVariables.appDriver = ''

        GlobalVariables.successPortal = "N/A"
        GlobalVariables.successApp = "N/A"


def pytest_sessionstart(session):
    print("Session setup level")
    server_logs.ssh_connection(router_ip, router_port, router_username, key_filename)


def pytest_sessionfinish(session, exitstatus):
    print("Session teardown level")
    updateExcel_With_Deselect_And_Broken()
    updateExcel_With_Category_And_Subcategory()
    updateExcel_With_RerunAttempts()
    setStylesForExcel()
    ssh.close()


def updateExcel_With_Deselect_And_Broken():
    wb = openpyxl.load_workbook(GlobalVariables.EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    for i in range(2, sheet.max_row + 1):
        colNum_testcase = ExcelProcessor.getColumnNumberFromName("", sheet, 'Test Case ID')
        testcase = (sheet.cell(row=i, column=colNum_testcase)).value
        if testcase in list_deselected_testcases:

            colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
            sheet.cell(row=i, column=colNum_overallStatus).value = "Deselected"

            colNum_execution = ExcelProcessor.getColumnNumberFromName("", sheet, 'TC Execution')
            sheet.cell(row=i, column=colNum_execution).value = "N/A"

            colNum_APIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'API Val')
            sheet.cell(row=i, column=colNum_APIval).value = "N/A"

            colNum_DBval = ExcelProcessor.getColumnNumberFromName("", sheet, 'DB Val')
            sheet.cell(row=i, column=colNum_DBval).value = "N/A"

            colNum_PortalVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'Portal Val')
            sheet.cell(row=i, column=colNum_PortalVal).value = "N/A"

            colNum_AppVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'App Val')
            sheet.cell(row=i, column=colNum_AppVal).value = "N/A"

            colNum_UIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'UI Val')
            sheet.cell(row=i, column=colNum_UIval).value = "N/A"

            colNum_exeTime = ExcelProcessor.getColumnNumberFromName("", sheet, 'Execution Time (sec)')
            sheet.cell(row=i, column=colNum_exeTime).value = "N/A"

            colNum_valTime = ExcelProcessor.getColumnNumberFromName("", sheet, 'Validation Time (sec)')
            sheet.cell(row=i, column=colNum_valTime).value = "N/A"

            colNum_logTime = ExcelProcessor.getColumnNumberFromName("", sheet, 'Log Coll Time (sec)')
            sheet.cell(row=i, column=colNum_logTime).value = "N/A"

            colNum_totTime = ExcelProcessor.getColumnNumberFromName("", sheet, 'Total Time (sec)')
            sheet.cell(row=i, column=colNum_totTime).value = "N/A"

        else:
            colNum_overallResult = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
            cellValue = (sheet.cell(row=i, column=colNum_overallResult)).value
            if cellValue is None:
                colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
                sheet.cell(row=i, column=colNum_overallStatus).value = "Broken"

                colNum_execution = ExcelProcessor.getColumnNumberFromName("", sheet, 'TC Execution')
                sheet.cell(row=i, column=colNum_execution).value = "N/A"

                colNum_APIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'API Val')
                sheet.cell(row=i, column=colNum_APIval).value = "N/A"

                colNum_DBval = ExcelProcessor.getColumnNumberFromName("", sheet, 'DB Val')
                sheet.cell(row=i, column=colNum_DBval).value = "N/A"

                colNum_PortalVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'Portal Val')
                sheet.cell(row=i, column=colNum_PortalVal).value = "N/A"

                colNum_AppVal = ExcelProcessor.getColumnNumberFromName("", sheet, 'App Val')
                sheet.cell(row=i, column=colNum_AppVal).value = "N/A"

                colNum_UIval = ExcelProcessor.getColumnNumberFromName("", sheet, 'UI Val')
                sheet.cell(row=i, column=colNum_UIval).value = "N/A"

                colNum_exeTime = ExcelProcessor.getColumnNumberFromName("", sheet, 'Execution Time (sec)')
                sheet.cell(row=i, column=colNum_exeTime).value = "N/A"

                colNum_valTime = ExcelProcessor.getColumnNumberFromName("", sheet, 'Validation Time (sec)')
                sheet.cell(row=i, column=colNum_valTime).value = "N/A"

                colNum_logTime = ExcelProcessor.getColumnNumberFromName("", sheet, 'Log Coll Time (sec)')
                sheet.cell(row=i, column=colNum_logTime).value = "N/A"

                colNum_totTime = ExcelProcessor.getColumnNumberFromName("", sheet, 'Total Time (sec)')
                sheet.cell(row=i, column=colNum_totTime).value = "N/A"

    wb.save(GlobalVariables.EXCEL_reportFilePath)


def updateExcel_With_Category_And_Subcategory():
    wb = openpyxl.load_workbook(GlobalVariables.EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    for i in range(2, sheet.max_row + 1):
        colNum_directoryName = ExcelProcessor.getColumnNumberFromName("", sheet, 'Directory Name')
        directoryName = (sheet.cell(row=i, column=colNum_directoryName)).value


        category = directoryName.split("/")[0]
        subCategory = directoryName.split("/")[1]

        colNum_category = ExcelProcessor.getColumnNumberFromName("", sheet, 'Category')
        colNum_subCategory = ExcelProcessor.getColumnNumberFromName("", sheet, 'Sub-Category')

        cellValue_category = (sheet.cell(row=i, column=colNum_category)).value
        cellValue_subCategory = (sheet.cell(row=i, column=colNum_subCategory)).value

        if cellValue_category is None:
            sheet.cell(row=i, column=colNum_category).value = category

            if cellValue_subCategory is None:
                sheet.cell(row=i, column=colNum_subCategory).value = subCategory



    wb.save(GlobalVariables.EXCEL_reportFilePath)


def updateExcel_With_RerunAttempts():
    wb = openpyxl.load_workbook(GlobalVariables.EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    if configReader.read_config("Validations", "rerun_immediately").lower() == "true" and configReader.read_config("Validations", "rerun_at_the_end").lower() == "false":
        colNum_rerunAttempts = ExcelProcessor.getColumnNumberFromName("", sheet, 'Rerun Attempts')
        for i in range(2, sheet.max_row + 1):
            colNum_testcase = ExcelProcessor.getColumnNumberFromName("", sheet, 'Test Case ID')
            testcase = (sheet.cell(row=i, column=colNum_testcase)).value

            rerunCount = Rerun.getRerunCount(testcase)
            if rerunCount >= 0:
                cellValue_rerunAttempts = int(configReader.read_config("Validations", "rerun_count")) - rerunCount
                sheet.cell(row=i, column=colNum_rerunAttempts).value = cellValue_rerunAttempts
            else:
                cellValue_rerunAttempts = configReader.read_config("Validations", "rerun_count")
                sheet.cell(row=i, column=colNum_rerunAttempts).value = int(cellValue_rerunAttempts)

    # If both rerun_at_the_end and rerun_immediately are disabled, setting value as N/A for Rerun Attempts
    if configReader.read_config("Validations", "rerun_at_the_end").lower() == "false" and configReader.read_config("Validations", "rerun_immediately").lower() == "false":
        colNum_RerunAttempts = ExcelProcessor.getColumnNumberFromName("", sheet, 'Rerun Attempts')
        for i in range(2, sheet.max_row + 1):
            sheet.cell(row=i, column=colNum_RerunAttempts).value = "N/A"

    wb.save(GlobalVariables.EXCEL_reportFilePath)


# ========================================================================================================================================

# def pytest_html_report_title(report):
#     report.title = 'Automation Test Detailed Report'
#
#
# def pytest_html_results_table_header(cells):
#     # removing old table headers
#     del cells[0]
#     del cells[1]
#     del cells[1]
#     # del cells[0]     #UNCOMMENT THIS TO REMOVE RERUN ITERATION COLUMN
#
#     # adding new headers
#     cells.insert(0, html.th('Status', class_='sortable time', col='time'))
#     cells.insert(1, html.th('TestCases'))
#     cells.insert(2, html.th('Duration (Seconds)'))
#     cells.insert(3, html.th('Rerun Iteration'))
#     cells.pop()
#
#
# #Once u override values of any column, you have to write code to override the later column values
# @pytest.mark.optionalhook
# def pytest_html_results_table_row(report, cells):
#     global rerunCount
#     change_html()
#
#     testcaseName = report.nodeid.split("::")[1]  # Extracting only the testcaseName from the format 'FolderName/Filename::testcaseName'
#
#     # duration = int(GlobalVariables.df_testCasesDetail['Total Time (sec)'][testcaseName]) # Fetching the duration of respective TC from the dataframe
#     duration = 100
#
#     # Populating TestCases column in main table
#     cells.insert(1, html.td(testcaseName))
#     cells.pop()
#
#     # Populating Duration column in main table
#     cells.insert(2, html.td(duration))
#     cells.pop()
#
#     if report.passed:
#         pass
#     elif report.failed:
#         pass
#     else:
#         del cells[:]   #Remove all the rows with status other than 'passed' and 'failed'
#
#     outcome = report.outcome
#
#     if outcome == 'rerun':
#         rerunCount = rerunCount + 1
#     else:
#         cells.insert(3, html.td(rerunCount))  # COMMENT THIS TO REMOVE RERUN ITERATION COLUMN
#         cells.pop()
#         rerunCount = 0
#
#
# @pytest.mark.hookwrapper
# @pytest.hookimpl(hookwrapper=True, tryfirst=True)
# def pytest_runtest_makereport(item, call):
#     global passed1
#     global failed1
#     global skipped1
#     global deselect1
#     pytest_html = item.config.pluginmanager.getplugin('html')
#     outcome = yield
#     report = outcome.get_result()
#     setattr(report, "duration_formatter", "%S")   #"%H:%M:%S.%f"
#     setattr(item, "rep_" + report.when, report)
#     report.description = str(item.function.__doc__)
#
#     if report.outcome == "passed" and report.when == "call":
#         passed1 = passed1 + 1
#
#     if report.outcome == "failed" and report.when == "call":
#         failed1 = failed1 + 1
#
#     if report.outcome == "skipped":
#         skipped1 = skipped1 + 1
#
#     return report
#
# @pytest.mark.optionalhook
# def pytest_html_results_summary(prefix, summary, postfix):
#     postfix.extend([html.style("table, th, td{border: 1px solid black;}")])
#
#
# def change_html():
#     global changeHTMLVal
#     global starting_time
#
#     now = datetime.now()
#     completed_time = now.strftime("%H:%M:%S")
#     FMT = '%H:%M:%S'
#     duration = datetime.strptime(completed_time, FMT) - datetime.strptime(str(starting_time), FMT)
#
#     if changeHTMLVal == 0:
#
#         env = configReader.read_config("APIs", "env")
#         f = open("template.css", 'r')
#         document = f.read()
#
#         document = document.replace("{bgcolor}", "linen")
#
#         # Data to the summary report table
#         skipped1 = GlobalVariables.testSuite_totalcases - (executed_tests_count + deselect1)
#
#         document = document.replace("{env}", env)
#         document = document.replace("{totalTC}", str(GlobalVariables.testSuite_totalcases))
#         document = document.replace("{passTC}", str(passed1))
#         # document = document.replace("{failTC}", str(tests_count-(passed1+skipped1)))    # FAILED TESTCASE NUMBER FROM CALCULATION
#         document = document.replace("{failTC}", str(failed1))   # FAILED TESTCASE NUMBER FROM GLOBAL VAR
#         document = document.replace("{skipTC}", str(skipped1))
#         document = document.replace("{deselectTC}", str(deselect1))
#
#         if passed1 != 0:
#             # perc = (passed1 / executed_tests_count) * 100
#             perc = (passed1 / GlobalVariables.testSuite_totalcases) * 100
#             document = document.replace("{pass%}", str(round(perc)) + "%")
#         else:
#             document = document.replace("{pass%}", "0%")
#
#         document = document.replace("{duration}", str(duration))
#
#         # Data for the pie chart
#         document = document.replace("{pie_passed}", str(passed1))
#         # document = document.replace("{pie_failed}", str(tests_count-(passed1+skipped1)))
#         document = document.replace("{pie_failed}", str(failed1)) # FAILED TESTCASE NUMBER FROM GLOBAL VAR
#         document = document.replace("{pie_skipped}", str(skipped1))
#         document = document.replace("{pie_deselect}", str(deselect1))
#
#         # Data to the validation failure count table
#         total_validationFailureCount = GlobalVariables.api_ValidationFailureCount + GlobalVariables.db_ValidationFailureCount + GlobalVariables.portal_ValidationFailureCount + GlobalVariables.app_ValidationFailureCount + GlobalVariables.ui_ValidationFailureCount + GlobalVariables.Incomplete_ExecutionCount
#
#         document = document.replace("{IncExe}", str(GlobalVariables.Incomplete_ExecutionCount))
#         document = document.replace("{apiCount}", str(GlobalVariables.api_ValidationFailureCount))
#         document = document.replace("{dbCount}", str(GlobalVariables.db_ValidationFailureCount))
#         document = document.replace("{portalCount}", str(GlobalVariables.portal_ValidationFailureCount))
#         document = document.replace("{appCount}", str(GlobalVariables.app_ValidationFailureCount))
#         document = document.replace("{UICount}", str(GlobalVariables.ui_ValidationFailureCount))
#         document = document.replace("{TotalValFail}", str(total_validationFailureCount))
#
#         file = open("test.html", "w")
#         file.write(document)
#         file.close()
#         changeHTMLVal = 1
#     else:
#         pass
#
#
# def pytest_runtestloop(session):
#     print("&&&&&&&&&&&&&&&")
#     global executed_tests_count
#     executed_tests_count = len(session.items) # TO GET TOTAL NUMBER OF EXECUTED TESTCASES
#     print(session.items)


# def pytest_configure(config):
#     config._metadata = None # To remove 'environment' table

# def createExcelWithHeaders(fileName):
#     path = '/home/ezetap/PycharmProjects/PortalAutomation/TestData/'+fileName+'.xlsx'
#     workbook = xlsxwriter.Workbook(path)
#     worksheet = workbook.add_worksheet('First')
#     format3 = workbook.add_format({'border': 1})
#     worksheet.write('A1', 'TestCases',format3)
#     worksheet.write('B1', 'TC Execution',format3)
#     worksheet.write('C1', 'API Val',format3)
#     worksheet.write('D1', 'DB Val',format3)
#     worksheet.write('E1', 'Portal Val',format3)
#     worksheet.write('F1', 'App Val',format3)
#     worksheet.write('G1', 'UI Val',format3)
#     worksheet.write('H1', 'Execution Time (sec)',format3)
#     worksheet.write('I1', 'Validation Time (sec)',format3)
#     worksheet.write('J1', 'Log Coll Time (sec)',format3)
#     worksheet.write('K1', 'Total Time (sec)',format3)
#     workbook.close()
#     return path


# def write_TC_Details_To_Excel():
#     TestCaseName = os.environ.get('PYTEST_CURRENT_TEST').replace(" (teardown)", '').replace("test_sample.py::",'').replace("TestCase/", '')
#     ExcelData = {"TC_Name": TestCaseName.replace("TestCase/",''),
#                  "TC_Execution": GlobalVariables.EXCEL_TC_Execution,
#                  "API_Val": GlobalVariables.EXCEL_API_Val,
#                  "DB_Val": GlobalVariables.EXCEL_DB_Val,
#                  "Portal_Val": GlobalVariables.EXCEL_Portal_Val,
#                  "App_Val": GlobalVariables.EXCEL_App_Val,
#                  "UI_Val": GlobalVariables.EXCEL_UI_Val,
#                  "Execution_Time": GlobalVariables.EXCEL_Execution_Time,
#                  "Val_time": GlobalVariables.EXCEL_Val_time,
#                  "LogCollTime": GlobalVariables.EXCEL_LogCollTime,
#                  "Tot_Time": GlobalVariables.EXCEL_Tot_Time}
#     # workbook = openpyxl.load_workbook(GlobalVariables.EXCEL_fileName) #Added the same in GlobalVariables file
#     setCellData(GlobalVariables.EXCEL_fileName, "First", GlobalVariables.EXCEL_rowNumber, 11, ExcelData)
#     GlobalVariables.EXCEL_rowNumber += 1


# def createDataframeForResult():
#     # For Dataframe used for testcase execution and result
#     dataForDataFrameHeader = {
#         'Test Case ID': [],
#         'File Name': [],
#         'TC Execution': [],
#         'API Val': [],
#         'DB Val': [],
#         'Portal Val': [],
#         'App Val': [],
#         'UI Val': [],
#         'Execution Time (sec)': [],
#         'Validation Time (sec)': [],
#         'Log Coll Time (sec)': [],
#         'Total Time (sec)': []
#     }
#
#     GlobalVariables.df_testCasesDetail = pd.DataFrame(dataForDataFrameHeader)
#
#     GlobalVariables.df_testCasesDetail.at[0, 'Test Case ID'] = "test_success"
#     GlobalVariables.df_testCasesDetail.at[1, 'Test Case ID'] = "test_Exe_Failure"
#     # GlobalVariables.df_testCasesDetail.at[2, 'Test Case ID'] = "test_api_val_exe_failure"
#     # GlobalVariables.df_testCasesDetail.at[3, 'Test Case ID'] = "test_DB_Val_Exe_Failure"
#     # GlobalVariables.df_testCasesDetail.at[4, 'Test Case ID'] = "test_portal_val_exe_failure"
#     # GlobalVariables.df_testCasesDetail.at[5, 'Test Case ID'] = "test_api_val_failure"
#     # GlobalVariables.df_testCasesDetail.at[6, 'Test Case ID'] = "test_DB_val_failure"
#     # GlobalVariables.df_testCasesDetail.at[7, 'Test Case ID'] = "test_portal_val_failure"
#
#     GlobalVariables.df_testCasesDetail.set_index('Test Case ID', inplace=True)
#
#
#     print(GlobalVariables.df_testCasesDetail.dtypes)
#
#     convert_dict = {'File Name': str,
#                     'TC Execution': str,
#                     'API Val': str,
#                     'DB Val': str,
#                     'Portal Val': str,
#                     'App Val': str,
#                     'UI Val': str,
#                     # 'Execution Time (sec)': int,
#                     # 'Validation Time (sec)': int,
#                     # 'Log Coll Time (sec)': int,
#                     # 'Total Time (sec)': int
#                     }
#
#
#     # GlobalVariables.df_testCasesDetail = GlobalVariables.df_testCasesDetail.astype(str) # Converting all columns to string
#
#     # GlobalVariables.df_testCasesDetail = GlobalVariables.df_testCasesDetail.astype(convert_dict)


# def pytest_sessionstart(session):
#     print("")
#     print("pytest_sessionstart")
#     session.results = dict()

# @pytest.mark.optionalhook
# def pytest_html_results_table_row(report, cells):
#     cells.insert(3, html.td("0"))
#     # cells.insert(1, html.td(datetime.utcnow(), class_="col-time")) #For time colulmn
#     # if report.passed:   #removes passed test cases
#     #     del cells[:]    #removes passed test cases
#     cells.pop()
#
#     # TO DELETE PASSED TC ROW
#     # if report.passed:
#     #     del cells[:]



#CONFIGURE ENVIRONMENT TABLE
# def pytest_configure(config):
    # username = getpwuid(getuid())[0]
    # py_version = python_version()
    # config._metadata = {
    #     "Org": username,
    #     "Python_version": py_version,
    #     "Environment":"dev3"
    # }
    # config._metadata = None #To remove 'environment' table



# def pytest_unconfigure(config):
#     print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
#     reporter = config.pluginmanager.get_plugin('terminalreporter')
#     duration = time.time() - reporter._sessionstarttime
#     reporter.write_sep('=', 'duration: {} seconds'.format(duration), yellow=True, bold=True)



#GET TOTAL TEST CASES EXECUTED
# def pytest_runtestloop(session):
#     global tests_count
#     tests_count = len(session.items)

# def pytest_terminal_summary(terminalreporter, exitstatus, config):
#     pass
    # duration = tm.time() - terminalreporter._sessionstarttime
    # duration = convert(duration)
    #
    # passedTC = len(terminalreporter.stats['passed'])
    # failedTC = len(terminalreporter.stats['failed'])
    # skippedTC = len(terminalreporter.stats['skipped'])
    #
    # f = open("11Febr2022.html", 'r')
    # document = f.read()
    #
    # document = document.replace("{passTC}", str(passedTC))
    # document = document.replace("{failTC}", str(failedTC))
    # document = document.replace("{skipTC}", str(skippedTC))
    # document = document.replace("{duration}", str(duration))
    # if(passedTC != 0):
    #     perc = (passedTC / tests_count) * 100
    #     document = document.replace("{pass%}", str(round(perc,2)) + "%")
    # else:
    #     document = document.replace("{pass%}", "0%")
    #
    # document = document.replace("{pie_passed}", str(passedTC))
    # document = document.replace("{pie_failed}", str(failedTC))
    # document = document.replace("{pie_skipped}", str(skippedTC))
    #
    # file = open("11Febr2022.html", "w")
    # file.write(document)
    # file.close()



#TO OPEN HTML REPORT AUTOMATICALLY : https://stackoverflow.com/questions/52032885/pytest-how-to-display-generated-report-after-test-run
# @pytest.hookimpl(trylast=True)
# def pytest_configure(config):
#     config._htmlfile = config._html.logfile



#https://python.plainenglish.io/create-your-customized-html-report-in-pytest-9c6b521b7e99
# @pytest.hookimpl(hookwrapper=True)
# def pytest_runtest_makereport(item, call):
#     '''data from the output of pytest gets processed here
#      and are passed to pytest_html_results_table_row'''
#     outcome = yield
#     # this is the output that is seen end of test case
#     report = outcome.get_result()
#     # taking doc string of the string
#     testcase = str(item.function.__doc__)
#     # name of the functton
#     c = str(item.function.__name__)[5:]
#
#     report.testcase = f"{c} [{testcase}]"
#     # taking input args
#     # example:
#     #      report.nodeid = 'tests/test_case.py::test_min[input0-1]'
#     #    data = re.split(r"\[|\]", 'tests/test_case.py::test_min[input0-1]')
#     #    =>  ['tests/test_case.py::test_min', 'input0-1', '']
#     report.tag = re.split(r"\[|\]", report.nodeid)[-2]

# # this is the output that is seen end of test case
    # report = outcome.get_result()
    # # taking doc string of the string
    # testcase = str(item.function.__doc__)
    # print(testcase)
    # # name of the functton
    # c = str(item.function.__name__)[5:]
    # print(c)
    # report.testcase = f"{c} [{testcase}]"
    # print(report.testcase) #TO GET ONLY TESTCASE NAME
    # print(report.nodeid) #TO GET FULL TEST CASE NAME


# def pytest_html_results_table_header(cells):
#     ''' meta programming to modify header of the result'''
#
#     from py.xml import html
#     # removing old table headers
#     del cells[1]
#     # adding new headers
#     cells.insert(0, html.th('Time', class_='sortable time', col='time'))
#     cells.insert(1, html.th('Tag'))
#     cells.insert(2, html.th('Testcase'))
#     cells.pop()
#
#
# def pytest_html_results_table_row(report, cells):
#     del cells[1]
#     cells.insert(0, html.td(datetime.utcnow(), class_='col-time'))
#     cells.insert(1, html.td(report.Tag))
#     cells.insert(2, html.td(report.testcase))
#     cells.pop()


# @pytest.hookimpl(trylast=True)
# def pytest_sessionfinish(session, exitstatus):
#     print("pytest_sessionfinish")
#
#     reporter = session.config.pluginmanager.get_plugin('terminalreporter')
#     duration = tm.time() - reporter._sessionstarttime
#     duration = convert(duration)
#
#     tests_count =  len(session.items)
#     passedTC = len(reporter.stats['passed'])
#     failedTC = len(reporter.stats['failed'])
#     skippedTC = len(reporter.stats['skipped'])
#
#     print(str(tests_count))
#     print(str(passedTC))
#     print(str(failedTC))
#     print(str(skippedTC))
#
#     env = configReader.read_config("APIs", "env")
#
#     f = open("template.css", 'r')
#     document = f.read()
#     time.sleep(5)
#
#     document = document.replace("{env}", env)
#     document = document.replace("{totalTC}", str(tests_count))
#     document = document.replace("{passTC}", str(passedTC))
#     document = document.replace("{failTC}", str(failedTC))
#     document = document.replace("{skipTC}", str(skippedTC))
#     document = document.replace("{duration}", str(duration))
#     if passedTC != 0:
#         perc = (passedTC / tests_count) * 100
#         document = document.replace("{pass%}", str(round(perc,2)) + "%")
#     else:
#         document = document.replace("{pass%}", "0%")
#
#     document = document.replace("{pie_passed}", str(passedTC))
#     document = document.replace("{pie_failed}", str(failedTC))
#     document = document.replace("{pie_skipped}", str(skippedTC))
#
#     file = open("test.html", "w")
#     time.sleep(5)
#     file.write(document)
#     time.sleep(5)
#     file.close()

# def convert(seconds):
#     print("convert")
#     seconds = seconds % (24 * 3600)
#     hour = seconds // 3600
#     seconds %= 3600
#     minutes = seconds // 60
#     seconds %= 60
#     return "%d:%02d:%02d" % (hour, minutes, seconds)



# @pytest.hookimpl(hookwrapper=True)
# def pytest_runtest_makereport(item, call):
#     '''data from the output of pytest gets processed here
#      and are passed to pytest_html_results_table_row'''
#     outcome = yield
#     # this is the output that is seen end of test case
#     report = outcome.get_result()
#     # taking doc string of the string
#     testcase = str(item.function.__doc__)
#     # name of the functton
#     c = str(item.function.__name__)[5:]
#
#     print(item.execution_count)              #TO GET EXECUTION COUNT OF EACH TEST CASE
#     print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
#     print(report.nodeid)
#     # report.tag = (report.testcase).split(r"\[|\]", report.nodeid)[-2]




# def pytest_html_results_table_row(report, cells):
#     ''' orienting the data gotten from  pytest_runtest_makereport
#     and sending it as row to the result '''
#     del cells[1]
#     cells.insert(0, html.td(datetime.utcnow(), class_='col-time'))
#     cells.insert(1, html.td(report.testcase))
#     cells.insert(2, html.td(report.testcase))
#     cells.pop()


# @pytest.mark.optionalhook
# def pytest_html_results_table_row(report, cells):
#     print("pytest_html_results_table_row")
#     change_html()
#     print("@@@@@@@@@@@@@@")
#     # print("@@@@@@@@@@@@@@")
#     # print("@@@@@@@@@@@@@@")
#     # print("")
#     # print(report)
#     # print(report.description)
#     print(report.nodeid)  #TO GET THE FULL TESTCASE
#
#
#     cells.insert(3, html.td("0"))  # COMMENT THIS TO REMOVE RERUN ITERATION COLUMN
#     cells.pop()




# def get_rerun_count(report):
#     global rerunCount
#     global currentTestCase
#     # testCase = report.nodeid   #TO GET TESTCASE NAME
#     # outcome = report.outcome