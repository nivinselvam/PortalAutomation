import os
import openpyxl
import pandas as pd

from DataProvider import GlobalVariables
from TestCase import ExcelProcessor, conftest
from Utilities import configReader

immediateRerun = True


def rerunTestAtTheEnd():
    df_rerunTestCases = pd.DataFrame(
        pd.read_excel(GlobalVariables.EXCEL_reportFilePath))

    df_rerunTestCases.set_index("Test Case ID", inplace=True)
    setOfRerunTest = set()

    exeFail = []
    apiValFail = []
    dbValFail = []
    portalValFail = []
    uiValFail = []
    appValFail = []
    setOfTest = set()

    for ind in df_rerunTestCases.index:
        testCaseName = str(df_rerunTestCases['File Name'][ind]) + ".py::" + str(ind)

        if str(df_rerunTestCases['TC Execution'][ind]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                      "exe_rerun") == "True":
            exeFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['API Val'][ind]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                 "api_rerun") == "True":
            apiValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['DB Val'][ind]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                "db_rerun") == "True":
            dbValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['Portal Val'][ind]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                    "portal_rerun") == "True":
            portalValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['App Val'][ind]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                 "app_rerun") == "True":
            appValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))
        if str(df_rerunTestCases['UI Val'][ind]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                "ui_rerun") == "True":
            uiValFail.append(testCaseName)
            setOfTest.add(testCaseName)
            setOfRerunTest.add(str(ind))



    ls_rerunTestCases = list(setOfTest)
    listToStr = ' '.join([str(elem) for elem in ls_rerunTestCases])

    if len(listToStr) > 0:

        # To send the testcaseID as a list to change the overall_Status as empty
        ls_TestCasesForRerun = list(setOfRerunTest)
        changeOverallStatusToEmpty(ls_TestCasesForRerun)

        # Executing the command for rerun at the end
        os.system(
            "pytest -v " + listToStr + ' --alluredir="/home/ezetap/PycharmProjects/PortalAutomation/TestCase/allure"')


def isRerunRequiredImmediately(testCaseID):
    isRerunRequired = False
    df_rerunTestCases = pd.DataFrame(
        pd.read_excel("/home/ezetap/PycharmProjects/PortalAutomation/TestCase/Report.xlsx"))

    df_rerunTestCases.set_index("Test Case ID", inplace=True)

    testCaseName = str(df_rerunTestCases['File Name'][testCaseID]) + ".py::" + str(testCaseID)

    if str(df_rerunTestCases['TC Execution'][testCaseID]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                         "exe_rerun") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['API Val'][testCaseID]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                    "api_rerun") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['DB Val'][testCaseID]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                   "db_rerun") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['Portal Val'][testCaseID]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                       "portal_rerun") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['App Val'][testCaseID]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                    "app_rerun") == "True":
        isRerunRequired = True
    if str(df_rerunTestCases['UI Val'][testCaseID]).lower() == 'fail' and configReader.read_config("Validations",
                                                                                                   "ui_rerun") == "True":
        isRerunRequired = True
    print("Rerun required = " + str(isRerunRequired))
    return isRerunRequired


def rerunTestImmediately(testCaseID, testCaseFileName, rerunCount, request):
    if setRerunCount(testCaseID, rerunCount):
        rerunCommand = "pytest -v " + testCaseFileName + ".py::" + testCaseID + ' --alluredir="/home/ezetap/PycharmProjects/PortalAutomation/TestCase/allure"'

        if rerunCount >=0:
            # To send the testcaseID as a list to change the overall_Status as empty
            setOfRerunTest = set()
            setOfRerunTest.add(testCaseID)
            ls_TestCasesForRerun = list(setOfRerunTest)
            changeOverallStatusToEmpty(ls_TestCasesForRerun)

            os.system(rerunCommand)
        if rerunCount == -1 and configReader.read_config("Validations", "rerun_at_the_end").lower() == "false":
            conftest.log_on_failure(request)

    else:
        print("Cannot perform rerun since the rerun count is 0 or the rerun sheet is not accessible.")


xl_RerunCountPath = "/home/ezetap/PycharmProjects/PortalAutomation/TestCase/RerunCount.xlsx"


# To change the value of rerun testcases overall_status to empty in Report excel, so that it will set as Broken in case of any connectivity issues
def changeOverallStatusToEmpty(ls_TestCasesForRerun):
    wb = openpyxl.load_workbook(GlobalVariables.EXCEL_reportFilePath)
    sheet = wb['Sheet1']

    for rerun_tesecase in ls_TestCasesForRerun:
        for i in range(2,sheet.max_row+1):
            colNum_testcase = ExcelProcessor.getColumnNumberFromName("", sheet, 'Test Case ID')
            testcase = (sheet.cell(row=i, column=colNum_testcase)).value

            if testcase == rerun_tesecase:
                colNum_overallStatus = ExcelProcessor.getColumnNumberFromName("", sheet, 'OverAll Results')
                sheet.cell(row=i, column=colNum_overallStatus).value = ""
    wb.save(GlobalVariables.EXCEL_reportFilePath)


def prepareImmediateRerunExcel():
    # df_overallTClist = pd.read_excel("/home/ezetap/PycharmProjects/PortalAutomation/DataProvider"
    #                                  "/TestCasesDetail.xlsx")

    df_overallTClist = pd.read_excel("/home/ezetap/PycharmProjects/PortalAutomation/TestCase/AllTestcaseSuite.xlsx")

    df_overallTClist.set_index('Test Case ID', inplace=True)
    df_overallTClist.drop(columns=['File Name', 'Execute','Unnamed: 0'], inplace=True)

    #Added for adding rerun attempts in Report excel
    df_overallTClist.drop(columns=['Directory Name'], inplace=True)
    reRunCount = []
    for i in range(0, len(df_overallTClist.index)):
        reRunCount.append(int(configReader.read_config("Validations", "rerun_count")))
        # reRunCount.append(GlobalVariables.int_immediateRerunCount)
    df_overallTClist['Rerun Count'] = reRunCount
    df_overallTClist.to_excel(xl_RerunCountPath, sheet_name="Rerun Count")

def prepareAtTheEndRerunExcel():
    df = pd.DataFrame({'Rerun Count': [int(configReader.read_config("Validations", "rerun_count"))]})
    writer = pd.ExcelWriter(xl_RerunCountPath, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Rerun At The End', index=False)
    writer.save()

def getRerunCount(testCaseID):
    try:
        df_TClist = pd.read_excel(xl_RerunCountPath, sheet_name="Rerun Count")
        df_TClist.set_index('Test Case ID', inplace=True)
        try:
            return df_TClist['Rerun Count'][testCaseID]
        except:
            None
    except:
        return -2


def getRerunCountForAtTheEnd():
    try:
        df_TClist = pd.read_excel(xl_RerunCountPath, sheet_name="Rerun At The End")
        # df_TClist.set_index('Test Case ID', inplace=True)
        try:
            return df_TClist['Rerun Count'][0]
        except:
            None
    except:
        return -2


def setRerunCount(testCaseID, rerunCount):
    try:
        workbook = openpyxl.load_workbook(xl_RerunCountPath)
        sheet = workbook["Rerun Count"]

        rowNumber = ExcelProcessor.getRowNumberFromValue(workbook, sheet, 'Test Case ID',
                                                         testCaseID)
        columnNumber = ExcelProcessor.getColumnNumberFromName(workbook, sheet, 'Rerun Count')
        sheet.cell(row=rowNumber, column=columnNumber).value = rerunCount

        workbook.save(xl_RerunCountPath)
        workbook.close()
        return True
    except:
        return False


def setRerunCountForAtTheEnd(rerunCount):
    try:
        workbook = openpyxl.load_workbook(xl_RerunCountPath)
        sheet = workbook["Rerun At The End"]

        rowNumber = 2
        columnNumber = 1
        sheet.cell(row=rowNumber, column=columnNumber).value = rerunCount

        workbook.save(xl_RerunCountPath)
        workbook.close()
        return True
    except:
        return False

# prepareImmediateRerunExcel()
