import logging
import time

import openpyxl
import pymysql
import pytest
import sshtunnel
from Utilities.Util_Logs import Logger
log = Logger(__name__, logging.INFO)

from selenium import webdriver
from selenium.webdriver.common.by import By
from self import self
import pandas as pd

from Pages.LoginPageActivities import LoginPageActivities
from Pages.HomePageActivities import HomePageActivities
from Pages.EditMerchantActivities import EditMerchantActivities
from Pages.UsersActivities import UsersActivities
from Pages.UserDetailsActivities import UserDetailsActivities

driver = webdriver.Chrome('/home/ezetap/Downloads/chromedriver_linux64/chromedriver')
portal_username = "8078151226"
portal_password = "D1234567"
username="1012101210"
password="A1234567"
orgCode="AUTO_PYTHON_TRIAL"
appkey = "appKey"

passwords=["!A1234567!/true","a1234567/true","A1234567/true","A12345/false","12345678/false","ABCDEFGH/false","abcdefgh/false"]

def setup_module(module):
    print("inside setup")
    LoginPageActivities.login(self, driver, portal_username, portal_password)
    time.sleep(5)
    HomePageActivities.search_for_merchant(self, driver, orgCode)
    HomePageActivities.switch_merchant(self, driver, orgCode)

def teardown_module(module):
    print("inside teardown")
    driver.close()

def updateToOldPassword():
    tunnel = sshtunnel.SSHTunnelForwarder(ssh_address_or_host='dev3', ssh_username="",
                                          remote_bind_address=('localhost', 3306))
    tunnel.start()
    db = pymysql.connect(host='localhost', user='ezedemo', passwd='abc123', db='', port=tunnel.local_bind_port)
    cursor = db.cursor()
    query = ("update ezetap_demo.org_employee set previous_passwords=NULL,password_hash ='vpdG0oSE5g6QIu0HYSET84sTvkpqH830JToCXIKGZZU=',failed_attempt_count='0' where username='1012101210';")
    cursor.execute(query)
    db.commit()
    print("updated PASSWORD")


def get_data(sheetName):
    workbook = openpyxl.load_workbook("/TestData/ChangePasswords.xlsx")
    sheet = workbook[sheetName]
    totalRows = sheet.max_row
    totalCols = sheet.max_column
    print("total cols are ", str(totalCols))
    print("total rows are ", str(totalRows))
    mainList = []

    for i in range(2, totalRows + 1):
        dataList = []
        for j in range(1, totalCols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
        print(mainList)
    return mainList

@pytest.mark.parametrize("password,success,expectedMessage,testcase", get_data("Sheet1"))
def test_change_pwd(password, success, expectedMessage,testcase):
    HomePageActivities.click_on_users_menu(self, driver)
    UsersActivities.click_edit_user(self, driver, username)
    UserDetailsActivities.enterPassword(self, driver, password)
    UserDetailsActivities.enterConfirmPassword(self, driver, password)
    UserDetailsActivities.save_details(self, driver)

    print("password : ",password)
    print("success : ",success)
    print("expectedMessage : ",expectedMessage)
    print("testcase : ",testcase)

    if success:
        print("SUCCESS IS TRUE")
        actualMessage = driver.find_element(By.XPATH, "//div[text()='" + expectedMessage + "']").text
        assert actualMessage == "User information is successfully updated.", ("Actual message:", actualMessage)
        updateToOldPassword()
    else:
        print("SUCCESS IS FALSE")
        actualMessage = driver.find_element(By.XPATH, "//span[text()='" + expectedMessage + "']").text
        assert actualMessage == expectedMessage, ("Actual message:", actualMessage)
